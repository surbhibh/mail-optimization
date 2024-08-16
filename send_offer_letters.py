import os
import smtplib
import time  # Import the time module for tracking execution time
from docxtpl import DocxTemplate
from docx2pdf import convert
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText
from openpyxl import load_workbook, Workbook
from datetime import datetime

# Load the Excel file
excel_file = 'employee_data.xlsx'  # Replace with your Excel file path
wb = load_workbook(excel_file)
ws = wb.active  # Assuming the data is in the first sheet

# Template file
template_file = 'OfferLetterTemplate.docx'  # Replace with your template path
current_directory = os.path.dirname(os.path.abspath(__file__))

# Create a folder named with today's date
today_date = datetime.now().strftime("%Y-%m-%d")
output_folder = os.path.join(current_directory, today_date)
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# Email setup (using Outlook SMTP)
SMTP_SERVER = 'smtp.office365.com'
SMTP_PORT = 587
EMAIL_USER = 'Your Email Address' # Replace with your Outlook email address
EMAIL_PASS = 'Your Password'       # Replace with your Outlook app password  

# Start tracking the total processing time
total_start_time = time.time()

def generate_pdf(agent_name, mobile_number, whatsapp_number, agent_mail_id, process, vendor_name, date_of_joining):
    """
    Generate a PDF for the offer letter.
    """
    # Add a timestamp or unique identifier to the file name
    unique_id = datetime.now().strftime("%Y%m%d%H%M%S")  # Current timestamp
    current_date = datetime.now().strftime("%d-%m-%Y")
    output_docx = os.path.join(output_folder, f"Offer_Letter_{agent_name}_{unique_id}.docx")
    
    context = {
        'Agent_Name': agent_name,
        'Agent_Mobile_Number': mobile_number,
        'Agent_Whatsapp_Number': whatsapp_number,
        'Agent_Email': agent_mail_id,
        'Process': process,
        'Vendor_Name': vendor_name,
        'Date_of_Joining': date_of_joining,  # New variable for Date of Joining
        'Todays_Date': current_date
        # Add more fields as required
    }
    
    # Generate the Word document
    doc = DocxTemplate(template_file)
    doc.render(context)
    doc.save(output_docx)
    
    # Convert the Word document to PDF
    output_pdf = output_docx.replace('.docx', '.pdf')
    try:
        convert(output_docx, output_pdf)
        print(f"Generated PDF for {agent_name} at {output_pdf}")
        return output_pdf
    except Exception as e:
        print(f"Failed to convert {output_docx} to PDF: {e}")
        return None

def send_email(receiver_email, subject, body, attachment_path):
    """
    Send an email with a PDF attachment.
    """
    msg = MIMEMultipart()
    msg['From'] = EMAIL_USER
    msg['To'] = receiver_email
    msg['Subject'] = subject
    
    msg.attach(MIMEText(body, 'plain'))

    with open(attachment_path, "rb") as attachment:
        part = MIMEApplication(attachment.read(), _subtype="pdf")
        part.add_header('Content-Disposition', 'attachment', filename=os.path.basename(attachment_path))
        msg.attach(part)
    
    try:
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()  # Secure the connection
        server.login(EMAIL_USER, EMAIL_PASS)
        server.sendmail(EMAIL_USER, receiver_email, msg.as_string())
        server.quit()
        print(f"Email sent to {receiver_email}")
        return True
    except Exception as e:
        print(f"Failed to send email to {receiver_email}: {e}")
        return False

# Process each row in the Excel file
total_records = ws.max_row - 1  # Total records excluding the header row
records_processed = 0

for index, row in enumerate(ws.iter_rows(min_row=2, values_only=False)):  # Skip the header row
    start_time = time.time()  # Track time for this employee

    offer_status = row[8].value  # Index 8 (ninth column) should be the "Offer Status" column

    if offer_status and offer_status.lower() == "sent":
        print(f"Skipping {row[2].value} as the offer has already been sent.")
        continue

    # In the provided code snippet, `row` is a variable that represents each row of
    # data being processed from the Excel file. It is used within a loop to iterate
    # over each row of the Excel sheet, extract specific values from that row (such as
    # email address, agent name, mobile number, etc.), and perform operations based on
    # the data in that row.
    email_address = row[1].value  # Index 1 (second column)
    agent_name = row[2].value  # Index 2 (third column)
    mobile_number = row[3].value
    whatsapp_number = row[4].value
    agent_mail_id = row[5].value
    process = row[6].value
    vendor_name = row[7].value
    date_of_joining = row[9].value  # Index 9 (10th column) for Date of Joining

    # Generate the PDF
    pdf_path = generate_pdf(agent_name, mobile_number, whatsapp_number, agent_mail_id, process, vendor_name, date_of_joining)
    
    if pdf_path:
        # Send the email with the generated PDF attached
        email_body = f"Dear {agent_name},\n\nPlease find your offer letter attached."
        email_sent = send_email(email_address, "Your Offer Letter", email_body, pdf_path)
        
        # Update the Excel sheet with the offer status
        if email_sent:
            ws.cell(row=index + 2, column=9).value = "Sent"  # Update the "Offer Status" column
        else:
            ws.cell(row=index + 2, column=9).value = "Failed"
    else:
        ws.cell(row=index + 2, column=9).value = "Failed"

    # Record the time taken for this employee
    time_taken = time.time() - start_time
    records_processed += 1
    print(f"Processed {agent_name} in {time_taken:.2f} seconds. ({records_processed}/{total_records} records processed)")

# Save the updated Excel sheet
wb.save(excel_file)

# Calculate and print the total time taken
total_time_taken = time.time() - total_start_time
print(f"All emails processed and offer statuses updated!")
print(f"Total time taken: {total_time_taken:.2f} seconds.")

